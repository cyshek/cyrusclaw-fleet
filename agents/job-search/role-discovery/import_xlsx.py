"""Import a Cyrus-edited Cyrus_Job_Tracker.xlsx back into tracker.db.

Only mutates Cyrus-editable fields:
  - status        (only when changed to 'skip' from blank)
  - applied_by
  - applied_on
  - cyrus_notes

Match key: Apply URL → source_key (via normalize_url). Falls back to
(company, role, loc) tuple match for noref: rows.

Dry-run by default. Pass --apply to actually write.

Usage:
    python import_xlsx.py [--xlsx PATH] [--apply]
"""
from __future__ import annotations
import argparse
import sys
from pathlib import Path

import openpyxl

sys.path.insert(0, str(Path(__file__).parent))
from tracker_db import connect, normalize_url

DEFAULT_XLSX = Path("/home/azureuser/.openclaw/agents/job-search/workspace/projects/job-search/Cyrus_Job_Tracker.xlsx")

EDITABLE = {"status", "applied_by", "applied_on", "cyrus_notes"}


def _norm(v):
    if v is None:
        return None
    if isinstance(v, str):
        s = v.strip()
        return s or None
    # openpyxl returns datetime for date cells
    try:
        return v.strftime("%Y-%m-%d")
    except Exception:
        return str(v).strip() or None


def parse_sheet(ws):
    headers = [c.value for c in ws[1]]
    # Map header label -> db field
    label_to_field = {
        "Company": "company",
        "Role": "role",
        "Location": "loc",
        "Apply URL": "app_url",
        "Cyrus notes": "cyrus_notes",
        "Applied by": "applied_by",
        "Applied on": "applied_on",
        "Status": "status",  # not currently exposed but tolerated
    }
    fields = []
    for h in headers:
        fields.append(label_to_field.get(h))
    rows = []
    present_fields = {f for f in fields if f}
    for excel_row in ws.iter_rows(min_row=2, values_only=True):
        rec = {"_present": present_fields}
        for f, v in zip(fields, excel_row):
            if f is None:
                continue
            rec[f] = _norm(v)
        if not rec.get("company"):
            continue
        rows.append(rec)
    return rows


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--xlsx", default=str(DEFAULT_XLSX))
    ap.add_argument("--apply", action="store_true", help="Write changes to DB (otherwise dry-run)")
    args = ap.parse_args()

    wb = openpyxl.load_workbook(args.xlsx, data_only=True)
    all_xlsx_rows = []
    for s in wb.sheetnames:
        all_xlsx_rows.extend(parse_sheet(wb[s]))
    print(f"Parsed {len(all_xlsx_rows)} data rows from {args.xlsx}")

    conn = connect()
    cur = conn.cursor()

    # Build lookups
    cur.execute("SELECT id, source_key, company, role, loc, status, applied_by, applied_on, cyrus_notes FROM roles")
    by_key = {}
    by_tuple = {}
    for r in cur.fetchall():
        d = dict(r)
        if d["source_key"]:
            by_key[d["source_key"]] = d
        by_tuple[(d["company"].strip().lower(), (d["role"] or "").strip().lower(), (d["loc"] or "").strip().lower())] = d

    proposed = []
    not_found = 0
    for x in all_xlsx_rows:
        key = normalize_url(x.get("app_url") or "")
        match = by_key.get(key) if key else None
        if match is None:
            t = (x["company"].strip().lower(),
                 (x.get("role") or "").strip().lower(),
                 (x.get("loc") or "").strip().lower())
            match = by_tuple.get(t)
        if match is None:
            not_found += 1
            continue
        diffs = {}
        present = x.get("_present", set())
        for f in EDITABLE:
            if f not in present:
                continue  # column not in this sheet → don't touch
            new = x.get(f)
            old = match.get(f)
            if (new or None) != (old or None):
                diffs[f] = (old, new)
        if diffs:
            proposed.append((match["id"], match["company"], match["role"], diffs))

    print(f"\nProposed updates: {len(proposed)}")
    print(f"XLSX rows with no DB match: {not_found}\n")
    for rid, co, role, diffs in proposed[:30]:
        print(f"  #{rid} {co} — {role}")
        for f, (old, new) in diffs.items():
            print(f"      {f}: {old!r} -> {new!r}")
    if len(proposed) > 30:
        print(f"  ... +{len(proposed)-30} more")

    if not args.apply:
        print("\n(Dry-run. Re-run with --apply to write.)")
        return

    for rid, _co, _role, diffs in proposed:
        sets = ", ".join(f"{f}=?" for f in diffs)
        vals = [new for (_old, new) in diffs.values()] + [rid]
        cur.execute(f"UPDATE roles SET {sets} WHERE id=?", vals)
    conn.commit()
    print(f"\nApplied {len(proposed)} updates.")


if __name__ == "__main__":
    main()
