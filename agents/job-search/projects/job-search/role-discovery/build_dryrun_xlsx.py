#!/usr/bin/env python3
"""Build a pretty XLSX summary of the Greenhouse dryrun specs."""
import json, glob, os
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule

HERE = Path(__file__).resolve().parent
DRYRUN = HERE.parent / "applications" / "dryrun"
OUT = HERE.parent / "outputs" / "greenhouse-dryrun-summary.xlsx"

rows = []
for f in sorted(glob.glob(str(DRYRUN / "*.json"))):
    d = json.load(open(f))
    c = d["counts"]
    blockers = "; ".join(b["label"][:60] for b in d.get("blockers", []))
    rows.append({
        "company": d.get("_company_name") or d["org"],
        "org": d["org"],
        "role": d.get("job_title") or "",
        "location": d.get("job_location") or "",
        "status": "READY" if d["ready_to_submit"] else f"BLOCKED ({c['blockers']})",
        "ready": d["ready_to_submit"],
        "filled": c["filled"],
        "needs_review": c["filled_needs_review"],
        "declined": c["declined"],
        "unresolved": c["unresolved"],
        "blockers_count": c["blockers"],
        "total": c["total_fields"],
        "blockers": blockers,
        "apply_url": d["role_url"],
    })

# Sort: READY first, then by company
rows.sort(key=lambda r: (not r["ready"], r["company"].lower(), r["role"].lower()))

wb = Workbook()
ws = wb.active
ws.title = "Greenhouse Queue"

headers = ["Status", "Company", "Role", "Location", "Filled", "Review", "Declined",
           "Unresolved", "Blockers", "Total", "Blocker Detail", "Apply URL"]
ws.append(headers)

# Header style
header_fill = PatternFill("solid", fgColor="1F2937")  # slate-800
header_font = Font(bold=True, color="FFFFFF", size=11, name="Calibri")
border_thin = Border(
    left=Side(style="thin", color="D1D5DB"),
    right=Side(style="thin", color="D1D5DB"),
    top=Side(style="thin", color="D1D5DB"),
    bottom=Side(style="thin", color="D1D5DB"),
)
for col_idx, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_idx)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = border_thin
ws.row_dimensions[1].height = 26

# Color fills for status
ready_fill = PatternFill("solid", fgColor="D1FAE5")     # emerald-100
ready_font = Font(color="065F46", bold=True)            # emerald-800
blocked_fill = PatternFill("solid", fgColor="FEE2E2")   # red-100
blocked_font = Font(color="991B1B", bold=True)          # red-800
alt_row_fill = PatternFill("solid", fgColor="F9FAFB")   # gray-50

# Body
for i, r in enumerate(rows, start=2):
    is_ready = r["ready"]
    ws.cell(row=i, column=1, value=r["status"])
    ws.cell(row=i, column=2, value=r["company"])
    ws.cell(row=i, column=3, value=r["role"])
    ws.cell(row=i, column=4, value=r["location"])
    ws.cell(row=i, column=5, value=r["filled"])
    ws.cell(row=i, column=6, value=r["needs_review"])
    ws.cell(row=i, column=7, value=r["declined"])
    ws.cell(row=i, column=8, value=r["unresolved"])
    ws.cell(row=i, column=9, value=r["blockers_count"])
    ws.cell(row=i, column=10, value=r["total"])
    ws.cell(row=i, column=11, value=r["blockers"])
    url_cell = ws.cell(row=i, column=12, value="Apply →")
    url_cell.hyperlink = r["apply_url"]
    url_cell.font = Font(color="2563EB", underline="single")

    # Status cell coloring
    sc = ws.cell(row=i, column=1)
    sc.fill = ready_fill if is_ready else blocked_fill
    sc.font = ready_font if is_ready else blocked_font
    sc.alignment = Alignment(horizontal="center", vertical="center")

    # Zebra rows for the rest
    if i % 2 == 0:
        for col in range(2, 13):
            c = ws.cell(row=i, column=col)
            if c.fill.fgColor.rgb in (None, "00000000"):
                c.fill = alt_row_fill

    # Borders + alignment everywhere
    for col in range(1, 13):
        c = ws.cell(row=i, column=col)
        c.border = border_thin
        if col in (5, 6, 7, 8, 9, 10):
            c.alignment = Alignment(horizontal="center", vertical="center")
        elif col == 11:
            c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        else:
            c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

# Column widths tuned by content
widths = {
    "A": 14,  # Status
    "B": 20,  # Company
    "C": 50,  # Role
    "D": 32,  # Location
    "E": 8,   # Filled
    "F": 8,   # Review
    "G": 9,   # Declined
    "H": 11,  # Unresolved
    "I": 10,  # Blockers
    "J": 7,   # Total
    "K": 50,  # Blocker detail
    "L": 12,  # Apply URL
}
for col, w in widths.items():
    ws.column_dimensions[col].width = w

# Freeze the header
ws.freeze_panes = "A2"

# Add an autofilter
ws.auto_filter.ref = f"A1:L{len(rows)+1}"

# Summary sheet
ws2 = wb.create_sheet("Summary")
ready_count = sum(1 for r in rows if r["ready"])
blocked_count = len(rows) - ready_count
summary = [
    ("Total roles in queue (Greenhouse)", len(rows)),
    ("READY to submit", ready_count),
    ("BLOCKED (need answers)", blocked_count),
    ("READY %", f"{ready_count*100//len(rows)}%" if rows else "0%"),
    ("Companies represented", len({r["company"] for r in rows})),
    ("Total form fields across all roles", sum(r["total"] for r in rows)),
    ("Fields auto-filled", sum(r["filled"] for r in rows)),
    ("Fields needing human input", sum(r["unresolved"] for r in rows)),
]
ws2.cell(row=1, column=1, value="Greenhouse dry-run summary").font = Font(bold=True, size=14, color="1F2937")
ws2.merge_cells("A1:B1")
for i, (k, v) in enumerate(summary, start=3):
    a = ws2.cell(row=i, column=1, value=k); a.font = Font(bold=True); a.fill = PatternFill("solid", fgColor="F3F4F6"); a.border = border_thin
    b = ws2.cell(row=i, column=2, value=v); b.border = border_thin; b.alignment = Alignment(horizontal="center")
ws2.column_dimensions["A"].width = 38
ws2.column_dimensions["B"].width = 16

OUT.parent.mkdir(parents=True, exist_ok=True)
wb.save(OUT)
print(f"Wrote {OUT}  ({OUT.stat().st_size} bytes, {len(rows)} rows)")
