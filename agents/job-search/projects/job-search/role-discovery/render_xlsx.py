"""Render tracker.db -> styled XLSX (modern minimalist).

Sheets (Cyrus 2026-06-20 — consolidated to 4):
  - Open         : roles not yet applied to and not skipped/closed/dead/blocked/manual.
  - Applied      : roles where Cyrus (or the agent) has submitted an application.
  - Manual Apply : EVERY non-auto-submitted role in ONE sheet — prepped-but-not-
                   submitted (prep_status='manual_ready', e.g. Workday packets),
                   hard-walled (status='blocked': OpenAI hold, Deepgram, captcha,
                   custom ATS), and apply-by-hand (status='manual-apply': Google
                   SSO, LinkedIn, Apple ID). A 'Why' column gives the short reason.
                   (Replaces the former separate "Manual Ready" + "Blocked" sheets.)
  - Interviews   : companies/roles where Cyrus received an interview (interviews table).
"""
from __future__ import annotations
import sys
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).parent))
from tracker_db import connect

OUT = Path("/home/azureuser/.openclaw/agents/job-search/workspace/projects/job-search/Cyrus_Job_Tracker.xlsx")

# Palette
HEADER_BG_OPEN    = "1F2937"  # slate-800
HEADER_BG_APPLIED = "065F46"  # emerald-800 — used for ALL headers on Applied sheet
HEADER_BG_MANUAL  = "7C2D12"  # amber-900 — "Cyrus to click Submit" sheet (Workday packets)
HEADER_BG_BLOCKED    = "7F1D1D"  # red-900 — blocked/walled rows (NOT actionable in Open triage)
HEADER_BG_INTERVIEWS = "1E3A5F"  # navy — Interviews sheet
HEADER_FG         = "FFFFFF"
LINK_FG           = "0563C1"
ZEBRA_BG          = "F8FAFC"
APPLIED_TINT      = "ECFDF5"
BORDER_COLOR      = "94A3B8"  # slate-400, visible but not heavy

HEADER_FONT  = Font(bold=True, color=HEADER_FG, size=11)
LINK_FONT    = Font(color=LINK_FG, underline="single", size=11)
BODY_FONT    = Font(size=11)
HEADER_ALIGN = Alignment(horizontal="left", vertical="center")
BODY_ALIGN   = Alignment(horizontal="left", vertical="center", wrap_text=False)

# Borders intentionally omitted — Cyrus prefers the borderless look.


def _pick_url(r: dict) -> str:
    return r.get("app_url") or r.get("jd_url") or ""


def _safe_link_url(u: str) -> str:
    """Return a clean http(s) URL safe to attach as an Excel hyperlink, else ''.

    Some DB rows carry malformed app_url values (e.g. a trailing backslash
    'https://employees-siriusxmradio.icims.com\\', or stray internal
    whitespace). render_xlsx used to attach these as hyperlinks anyway because
    they start with 'http', producing a cell Cyrus can click that goes nowhere.
    Here we validate scheme + netloc and reject obvious corruption so the cell
    falls back to PLAIN TEXT (still readable) instead of a broken link.
    This is a pure-rendering guard — it never mutates the DB; it only decides
    whether a given value is link-worthy for THIS sheet.
    """
    if not u or not isinstance(u, str):
        return ""
    s = u.strip()
    # Internal whitespace or a trailing backslash => corrupted, not linkable.
    if any(ch.isspace() for ch in s) or s.endswith("\\"):
        return ""
    try:
        from urllib.parse import urlparse
        p = urlparse(s)
    except Exception:
        return ""
    if p.scheme not in ("http", "https"):
        return ""
    # netloc must exist and contain a dot (a real host, not a bare token).
    if not p.netloc or "." not in p.netloc:
        return ""
    return s


def _fmt_tc(r: dict) -> str:
    """Render est_tc as $NNNK (or blank if unknown). Blank = unknown = NEUTRAL."""
    v = r.get("est_tc")
    if not v:
        return ""
    try:
        return f"${int(v)//1000}K"
    except (TypeError, ValueError):
        return ""


COMP_FLOOR = 180_000  # Above this = prioritized. Below or unknown = unchanged (never dropped).


class _NegStr:
    """Wrap a string so it sorts in DESCENDING order under an ascending sort.

    Used to get FRESHEST-FIRST (posted_on/first_seen DESC) ordering inside an
    otherwise-ascending tuple sort key, without reversing the whole sort.
    Empty string compares as 'oldest' (sorts last among newest-first).
    """
    __slots__ = ("s",)

    def __init__(self, s: str):
        self.s = s or ""

    def __lt__(self, other: "_NegStr") -> bool:
        # Reversed comparison => larger (newer) string is "less" => sorts first.
        return self.s > other.s

    def __eq__(self, other) -> bool:
        return isinstance(other, _NegStr) and self.s == other.s


# Title tokens that strongly imply senior+ regardless of stated YOE.
_SENIOR_TITLE_RE = __import__("re").compile(
    r"\b(senior|sr\.?|staff|principal|lead|director|head of|vp|chief|distinguished|fellow|group|partner|architect ii+|architect iii+|level 3|level iii|l[4-9]|l1[0-9])\b",
    __import__("re").IGNORECASE,
)


def _fmt_experience(r: dict) -> str:
    """Best-available YOE/seniority hint, preferring the LLM-extracted value.

    Priority: llm_yoe_required (number) > exp_req-non-unstated > llm_seniority > title-derived senior flag > 'unstated'.
    """
    yoe = r.get("llm_yoe_required")
    if yoe not in (None, "", "unstated"):
        try:
            n = int(yoe)
            return f"{n}+ yrs (LLM)"
        except (TypeError, ValueError):
            return f"{yoe} (LLM)"
    exp = (r.get("exp_req") or "").replace("exp:", "")
    if exp and exp != "unstated":
        return exp
    sen = (r.get("llm_seniority") or "").strip()
    if sen and sen.lower() not in ("unstated", "unknown", "none"):
        return f"{sen} (LLM)"
    title = r.get("role") or ""
    if _SENIOR_TITLE_RE.search(title):
        return "senior? (title)"
    return "unstated"


def _recency_key(r: dict) -> str:
    """Freshest-first recency string for tie-breaking within a company group.

    Cyrus 2026-06-08: the sheet's Google section should be ordered
    FRESHEST-FIRST. Google JD pages expose no reliable posting date, so we
    order by posted_on DESC, falling back to first_seen DESC where posted_on
    is empty. Returning a string we then sort DESC (reverse) gives newest
    first; empty dates sort last. This is generic (helps every company's
    section), but it's the wiring that satisfies the Google recency ask.
    """
    return (r.get("posted_on") or r.get("first_seen") or "")


def _open_sort_key(r: dict):
    """Sort: above-floor first (highest TC), then unknowns, then below-floor.

    Unknowns are NEVER dropped — they sort between above-floor and below-floor.
    Within the same TC bucket + company, rows are ordered FRESHEST-FIRST
    (posted_on DESC, then first_seen DESC) — Cyrus 2026-06-08. We negate the
    recency string by sorting on its codepoints reversed so the OUTER ascending
    sort still yields newest-first inside each (bucket, company).
    """
    tc = r.get("est_tc")
    company = (r.get("company") or "").lower()
    role = (r.get("role") or "").lower()
    # _NegStr makes a string compare in DESCENDING order under an ascending sort
    # (so newest ISO date sorts first). Empty -> sorts last (oldest).
    recency = _NegStr(_recency_key(r))
    if tc and tc >= COMP_FLOOR:
        return (0, -int(tc), company, recency, role)
    if not tc:
        return (1, 0, company, recency, role)
    return (2, -int(tc), company, recency, role)


def _format_response(r: dict) -> str:
    """Combine last_response_at + response_status for the 'Last response' col."""
    when = r.get("last_response_at") or ""
    status = r.get("response_status") or ""
    if when and status:
        return f"{when} · {status}"
    return when or status or ""


def _is_workday_url(u: str) -> bool:
    """True if the apply URL points at a Workday tenant (any wdN host)."""
    s = (u or "").lower()
    return (
        "myworkdayjobs" in s
        or "workday" in s
        or any(f"wd{n}." in s for n in range(1, 13))
    )


def _why_label(r: dict) -> str:
    """Human-readable short 'Why' label for the consolidated Manual Apply sheet.

    The consolidated sheet folds together THREE prior buckets — prepped-but-not-
    submitted (prep_status='manual_ready'), hard-walled (status='blocked'), and
    apply-by-hand (status='manual-apply'). block_reason values in the DB range
    from tidy slugs ('openai-applimit-180d') to multi-line audit essays, so this
    maps by PREFIX/substring to a stable short label. Order matters: the most
    specific / highest-signal checks come first. Falls back to a truncated
    block_reason (or a prep/url-derived label) so nothing is ever unlabeled.
    """
    br = (r.get("block_reason") or "").strip()
    brl = br.lower()
    status = (r.get("status") or "").lower()
    prep = (r.get("prep_status") or "").lower()
    url = r.get("app_url") or r.get("jd_url") or ""

    # --- highest-signal known walls (substring/prefix tolerant) ---
    if "openai-applimit" in brl or "openai-180" in brl:
        return "OpenAI 180-day hold"
    if "deepgram-email-blocked" in brl or brl.startswith("deepgram"):
        return "Deepgram email-blocked until ~Jul 30"
    if "google-sso" in brl:
        return "Google SSO (Cyrus)"
    if "linkedin-no-external-apply" in brl:
        return "LinkedIn (no external apply)"
    if "linkedin-no-ats-found" in brl:
        return "LinkedIn (no ATS found)"
    if "linkedin" in brl:
        # LinkedIn-auth-stranded / li_at-unusable / stranded variants.
        return "LinkedIn (auth-stranded)"
    if "lever-hcaptcha" in brl or ("lever" in brl and "hcaptcha" in brl):
        return "Lever hCaptcha (need nopecha)"
    if "ashby-hard-recaptcha" in brl or "ashby-hard-score" in brl \
            or "ashby-score-gate" in brl or "recaptcha_score_below_threshold" in brl \
            or "recaptcha-spam-flag" in brl:
        return "Ashby reCAPTCHA (need residential IP)"
    if "ashby-knockout" in brl:
        return "Ashby knockout question"
    if "icims" in brl and "hcaptcha" in brl:
        return "iCIMS hCaptcha (no solver)"
    if "icims" in brl:
        return "iCIMS (account/req gate)"
    if "bytedance" in brl or "tiktok" in brl:
        return "ByteDance/TikTok OTP-captcha"
    if "apple-id" in brl or "need-runner-apple" in brl:
        return "Apple ID SSO (Cyrus)"
    if "company-blocklist" in brl:
        return "Blocklisted (Cyrus handles)"
    if "senior-title" in brl:
        return "Senior-title out of scope"
    if "sf-bay-area" in brl or "sf-office" in brl:
        return "SF Bay Area required"
    if brl.startswith("need-runner-") or "-no-runner" in brl or "no-runner" in brl:
        return "Custom ATS (no runner)"
    if brl.startswith("gh-custom-q") or "gh-uncertain" in brl:
        return "Greenhouse custom Q (unanswered)"
    if "workday" in brl:
        # workday dupe-class / how-did-you-hear / fuzzy-match notes.
        return "Workday (manual verify)"
    if "teamtailor" in brl:
        return "Teamtailor (no runner)"
    if "jobdiva" in brl or "ripplehire" in brl or "contactrh" in brl \
            or "oracle-hcm" in brl or "amazon-custom" in brl or "snap-no-public" in brl:
        return "Custom ATS (no runner)"
    if "manual-apply-only" in brl or "manual-apply" in brl:
        return "Manual apply only"

    # --- no explicit block_reason: derive from prep_status / url ---
    if not br:
        if prep == "manual_ready":
            if _is_workday_url(url):
                return "Workday (prepped, click Submit)"
            return "Prepped, needs runner attempt"
        if status == "manual-apply":
            return "Manual apply (by hand)"
        if status in ("blocked", "scan-blocked"):
            return "Blocked (see notes)"
        return ""

    # --- fallback: truncated raw block_reason (first line, capped) ---
    first_line = br.splitlines()[0]
    return first_line if len(first_line) <= 60 else first_line[:57] + "…"


def _manual_apply_sort_key(r: dict):
    """Sort the consolidated Manual Apply sheet.

    Cyrus 2026-06-20: Workday-prepped rows (actionable — he just clicks Submit)
    sort FIRST, then everything else by company alpha. Within Workday-prepped,
    also company alpha then role.
    """
    prep = (r.get("prep_status") or "").lower()
    url = r.get("app_url") or r.get("jd_url") or ""
    is_wd_prepped = 0 if (prep == "manual_ready" and _is_workday_url(url)) else 1
    company = (r.get("company") or "").lower()
    role = (r.get("role") or "").lower()
    return (is_wd_prepped, company, role)


def write_sheet(ws, rows, columns, header_bg, zebra_fill):
    header_fill = PatternFill("solid", fgColor=header_bg)
    for col_idx, (field, label, _w) in enumerate(columns, 1):
        c = ws.cell(row=1, column=col_idx, value=label)
        c.font = HEADER_FONT
        c.fill = header_fill
        c.alignment = HEADER_ALIGN

    for row_idx, r in enumerate(rows, start=2):
        is_zebra = (row_idx % 2 == 0)
        for col_idx, (field, _label, _w) in enumerate(columns, 1):
            if field == "url":
                val = _pick_url(r)
            elif field == "why":
                val = _why_label(r)
            elif field == "response":
                val = _format_response(r)
            elif field == "est_tc":
                val = _fmt_tc(r)
            elif field == "exp_req":
                val = _fmt_experience(r)
            elif field == "source_key":
                sk = r.get("source_key") or ""
                val = sk.split(":")[0] if ":" in sk else sk
            else:
                val = r.get(field)
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.alignment = BODY_ALIGN
            cell.font = BODY_FONT
            if False and is_zebra:
                cell.fill = zebra_fill
            if field == "url" and val and isinstance(val, str):
                link = _safe_link_url(val)
                if link:
                    cell.hyperlink = link
                    cell.value = link if len(link) <= 60 else link[:57] + "…"
                    cell.font = LINK_FONT

    for col_idx, (_f, _l, w) in enumerate(columns, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    ws.sheet_view.showGridLines = True


def build():
    conn = connect()
    cur = conn.cursor()
    cur.execute("SELECT * FROM roles ORDER BY company COLLATE NOCASE, role COLLATE NOCASE")
    all_rows = [dict(r) for r in cur.fetchall()]

    open_rows = [
        r for r in all_rows
        if not r["applied_by"]
        and (r.get("prep_status") or "") not in ("manual_ready",)
        and (r["status"] or "") not in ("skip", "closed", "none", "scan-blocked", "blocked", "manual-apply")
    ]
    open_rows.sort(key=_open_sort_key)
    applied_rows = [r for r in all_rows if r["applied_by"]]
    applied_rows.sort(key=lambda r: (r["applied_on"] or "", r["company"].lower()), reverse=True)
    # ── CONSOLIDATED "Manual Apply" set (Cyrus 2026-06-20) ──────────────────────
    # ONE sheet for every role that is NOT auto-submitted — i.e. every role that
    # genuinely exhausted (or can't use) the automated path. Folds together three
    # former buckets:
    #   (a) prep_status='manual_ready'  — prepped but not submitted (Workday packets
    #       Cyrus clicks Submit on; captcha/score-walled Ashby/Lever prepped rows).
    #   (b) status='blocked' / 'scan-blocked' — hard walls (OpenAI 180-day hold,
    #       Deepgram email-block, Ashby HARD reCAPTCHA, Lever hCaptcha, custom ATS).
    #   (c) status='manual-apply' — apply-by-hand (Google SSO, LinkedIn, Apple ID,
    #       custom ATSes with no runner).
    # A 'Why' column (via _why_label) gives each row a short human reason. Workday-
    # prepped rows sort FIRST (actionable: just click Submit), then company alpha.
    # The old separate "Manual Ready" and "Blocked" sheets are REMOVED.
    seen_ids = set()
    consolidated_rows = []
    for r in all_rows:
        if r["applied_by"]:
            continue
        prep = (r.get("prep_status") or "")
        status = (r["status"] or "")
        include = (
            prep == "manual_ready"
            or status in ("blocked", "scan-blocked", "manual-apply")
        )
        if not include:
            continue
        if r["id"] in seen_ids:
            continue
        seen_ids.add(r["id"])
        consolidated_rows.append(r)
    consolidated_rows.sort(key=_manual_apply_sort_key)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    ws_open = wb.create_sheet(f"Open ({len(open_rows)})")
    open_cols = [
        ("company",     "Company",     20),
        ("role",         "Role",        48),
        ("loc",          "Location",    24),
        ("exp_req",      "Experience",  13),
        ("est_tc",       "Est TC",       9),
        ("posted_on",    "Posted",      12),
        ("first_seen",   "First seen",  12),
        ("status",       "Status",      14),
        ("url",          "Apply URL",   55),
        ("applied_on",   "Applied on",  16),
        ("cyrus_notes",  "Cyrus notes", 46),
        ("agent_notes",  "Agent notes", 46),
    ]
    write_sheet(ws_open, open_rows, open_cols,
                header_bg=HEADER_BG_OPEN, zebra_fill=PatternFill("solid", fgColor=ZEBRA_BG))

    ws_app = wb.create_sheet(f"Applied ({len(applied_rows)})")
    applied_cols = [
        ("company",          "Company",        20),
        ("role",              "Role",           48),
        ("loc",               "Location",       24),
        ("posted_on",         "Posted",         12),
        ("applied_on",        "Applied on",     16),
        ("applied_by",        "Applied by",     16),
        ("response",          "Last response",  22),
        ("last_email_from",   "Email from",     32),
        ("last_email_subject","Subject",        50),
        ("cyrus_notes",       "Cyrus notes",    46),
        ("agent_notes",       "Agent notes",    46),
        ("url",               "Apply URL",      50),
    ]
    write_sheet(ws_app, applied_rows, applied_cols,
                header_bg=HEADER_BG_APPLIED, zebra_fill=PatternFill("solid", fgColor=APPLIED_TINT))

    # ── Manual Apply (CONSOLIDATED) ───────────────────────────────────
    # ONE sheet for every non-auto-submitted role (prepped-not-submitted + blocked +
    # apply-by-hand). The 'Why' column carries a short human reason (_why_label).
    # Workday-prepped rows sort first (actionable). Replaces the old separate
    # "Manual Ready" and "Blocked" sheets, which are now REMOVED.
    ws_ma = wb.create_sheet(f"Manual Apply ({len(consolidated_rows)})")
    manual_apply_cols = [
        ("company",      "Company",        20),
        ("role",         "Role",           48),
        ("why",          "Why",            34),
        ("loc",          "Location",       22),
        ("exp_req",      "Experience",     13),
        ("posted_on",    "Posted",         12),
        ("url",          "Apply URL",      55),
        ("source_key",   "Source",         16),
        ("prep_path",    "Packet path",    50),
        ("cyrus_notes",  "Cyrus notes",    40),
        ("agent_notes",  "Agent notes",    40),
    ]
    write_sheet(ws_ma, consolidated_rows, manual_apply_cols,
                header_bg=HEADER_BG_MANUAL, zebra_fill=PatternFill("solid", fgColor=ZEBRA_BG))

    # Interviews: companies where Cyrus received an interview invitation.
    cur = conn.execute("""
        SELECT i.id, i.company, i.role, i.jd_url, i.applied_on,
               i.interview_type, i.interview_date, i.outcome, i.notes, i.added_on
        FROM interviews i
        ORDER BY i.company, i.role
    """)
    cols_iv = [d[0] for d in cur.description]
    interview_rows = [dict(zip(cols_iv, row)) for row in cur.fetchall()]
    # Normalise so write_sheet can pick up 'url'
    for r in interview_rows:
        r["url"] = r.get("jd_url") or ""

    ws_iv = wb.create_sheet(f"Interviews ({len(interview_rows)})")
    interview_cols = [
        ("company",        "Company",         22),
        ("role",           "Role",             50),
        ("applied_on",     "Applied on",       14),
        ("interview_type", "Interview type",   20),
        ("interview_date", "Interview date",   16),
        ("outcome",        "Outcome",          18),
        ("notes",          "Notes",            50),
        ("url",            "JD URL",           55),
    ]
    write_sheet(ws_iv, interview_rows, interview_cols,
                header_bg=HEADER_BG_INTERVIEWS, zebra_fill=PatternFill("solid", fgColor=ZEBRA_BG))

    conn.close()
    wb.save(OUT)
    print(f"Wrote: {OUT}")
    print(f"  Open:         {len(open_rows)}")
    print(f"  Applied:      {len(applied_rows)}")
    print(f"  Manual Apply: {len(consolidated_rows)}")
    print(f"  Interviews:   {len(interview_rows)}")


if __name__ == "__main__":
    build()
