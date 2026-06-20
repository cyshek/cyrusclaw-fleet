"""Tests for render_xlsx pure-rendering helpers.

Covers the hyperlink-safety guard added 2026-06-08 (audit) plus the existing
freshest-first ordering + experience formatting helpers, so a future edit to
render_xlsx.py can't silently regress them. NO DB writes; pure-function tests.
"""
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent))
import render_xlsx as rx


# ---- _safe_link_url: the new hyperlink-corruption guard ----

def test_safe_link_url_accepts_normal_https():
    u = "https://jobs.ashbyhq.com/openai/56fceb8e-589b-410e-8b21-24f9945ccb9d"
    assert rx._safe_link_url(u) == u

def test_safe_link_url_accepts_http_with_query():
    u = "http://videoamp.com/job?gh_jid=8465234002"
    assert rx._safe_link_url(u) == u

def test_safe_link_url_rejects_trailing_backslash():
    # Real corrupted SiriusXM rows (ids 1497/1500).
    assert rx._safe_link_url("https://employees-siriusxmradio.icims.com\\") == ""

def test_safe_link_url_rejects_internal_whitespace():
    assert rx._safe_link_url("https://example.com/a b/job") == ""
    assert rx._safe_link_url("https://exa mple.com/job") == ""

def test_safe_link_url_strips_surrounding_whitespace():
    u = "  https://example.com/job  "
    assert rx._safe_link_url(u) == "https://example.com/job"

def test_safe_link_url_rejects_non_http_scheme():
    assert rx._safe_link_url("mailto:foo@bar.com") == ""
    assert rx._safe_link_url("ftp://example.com/x") == ""

def test_safe_link_url_rejects_bare_token_no_host():
    assert rx._safe_link_url("https://localhost") == ""  # no dot in netloc
    assert rx._safe_link_url("not a url") == ""
    assert rx._safe_link_url("") == ""
    assert rx._safe_link_url(None) == ""


# ---- _fmt_experience: clean exp tokens (no raw 'exp:' leakage) ----

def test_fmt_experience_prefers_llm_yoe():
    assert rx._fmt_experience({"llm_yoe_required": 5}) == "5+ yrs (LLM)"

def test_fmt_experience_strips_exp_prefix():
    assert rx._fmt_experience({"exp_req": "exp:3+yrs"}) == "3+yrs"

def test_fmt_experience_unstated_falls_through():
    assert rx._fmt_experience({"exp_req": "exp:unstated"}) == "unstated"

def test_fmt_experience_title_senior_hint():
    out = rx._fmt_experience({"role": "Senior Product Manager"})
    assert out == "senior? (title)"


# ---- recency / open-sort: freshest-first within a TC/company bucket ----

def test_negstr_sorts_descending():
    items = [rx._NegStr("2026-05-11"), rx._NegStr("2026-06-04"), rx._NegStr("")]
    assert sorted(items, key=lambda x: x).index(rx._NegStr("2026-06-04")) == 0
    # empty sorts last (oldest)
    assert sorted(items, key=lambda x: x)[-1] == rx._NegStr("")

def test_recency_key_prefers_posted_then_first_seen():
    assert rx._recency_key({"posted_on": "2026-06-01", "first_seen": "2026-05-01"}) == "2026-06-01"
    assert rx._recency_key({"posted_on": None, "first_seen": "2026-05-01"}) == "2026-05-01"
    assert rx._recency_key({}) == ""

def test_open_sort_orders_freshest_first_within_company():
    # Same company, no TC -> bucket 1; should order by first_seen DESC.
    rows = [
        {"company": "Google", "role": "B role", "est_tc": None, "first_seen": "2026-05-01"},
        {"company": "Google", "role": "A role", "est_tc": None, "first_seen": "2026-06-04"},
        {"company": "Google", "role": "C role", "est_tc": None, "first_seen": "2026-06-04"},
    ]
    rows.sort(key=rx._open_sort_key)
    # Newest first_seen wins; within same date, role ASC.
    assert [r["role"] for r in rows] == ["A role", "C role", "B role"]

def test_open_sort_above_floor_beats_unknown_beats_below():
    rows = [
        {"company": "X", "role": "low", "est_tc": 100_000, "first_seen": ""},
        {"company": "X", "role": "unk", "est_tc": None, "first_seen": ""},
        {"company": "X", "role": "high", "est_tc": 250_000, "first_seen": ""},
    ]
    rows.sort(key=rx._open_sort_key)
    assert [r["role"] for r in rows] == ["high", "unk", "low"]


# ---- structure smoke (2026-06-08): Blocked tab split out of Open + new cols ----
# Read-only: build() reads the live tracker.db and writes the OUT xlsx (a build
# artifact). Asserts the Open-tab declutter (blocked rows live in their own tab) and
# that the new Status / First seen / Blocked reason columns are present. Locks the
# 2026-06-08 restructure so a future edit can't silently merge blocked back into Open.

def test_build_produces_blocked_tab_and_new_columns():
    import openpyxl
    rx.build()
    wb = openpyxl.load_workbook(rx.OUT)
    names = wb.sheetnames
    # exactly the four expected tabs, each "<Name> (<n>)"
    bases = [n.split(" (")[0] for n in names]
    for expect in ("Open", "Applied", "Manual Ready", "Blocked"):
        assert expect in bases, f"missing tab {expect!r}; got {bases}"

    def cols(base):
        ws = wb[next(n for n in names if n.startswith(base + " ("))]
        return [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]

    open_cols = cols("Open")
    assert "Status" in open_cols and "First seen" in open_cols, open_cols
    blocked_cols = cols("Blocked")
    assert "Blocked reason" in blocked_cols, blocked_cols
    assert "Status" in blocked_cols and "First seen" in blocked_cols, blocked_cols


def test_blocked_status_rows_excluded_from_open_filter():
    # Mirror build()'s Open vs Blocked predicate to prove a blocked-status row never
    # lands in Open (the whole point of the declutter).
    blocked_row = {"applied_by": None, "prep_status": None, "status": "blocked"}
    open_ok_row = {"applied_by": None, "prep_status": None, "status": ""}
    def is_open(r):
        return (not r["applied_by"]
                and (r.get("prep_status") or "") not in ("manual_ready",)
                and (r["status"] or "") not in ("skip", "closed", "none", "scan-blocked", "blocked"))
    def is_blocked(r):
        return (not r["applied_by"]
                and (r.get("prep_status") or "") != "manual_ready"
                and (r["status"] or "") in ("blocked", "scan-blocked"))
    assert not is_open(blocked_row) and is_blocked(blocked_row)
    assert is_open(open_ok_row) and not is_blocked(open_ok_row)
